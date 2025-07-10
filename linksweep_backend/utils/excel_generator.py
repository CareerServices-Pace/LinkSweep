from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from zoneinfo import ZoneInfo
from datetime import datetime
import io, os, time

os.environ["TZ"] = "America/New_York"
time.tzset()

def generate_excel_report(broken_links: list[dict], scan_id: int):
    wb = Workbook()
    ws = wb.active
    ws.title = "Broken Links Report"

    # Define header and styles
    headers = ["Source Page", "Link", "Status", "Link Type", "Fix Guide"]

    # Styles
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Write headers with styles
    for col_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = gray_fill
        cell.alignment = center_alignment
        cell.border = border

    # Write data rows
    for row_index, item in enumerate(broken_links, 2):
        ws.cell(row=row_index, column=1, value=item["sourcePage"])
        ws.cell(row=row_index, column=2, value=item["link"])
        ws.cell(row=row_index, column=3, value=str(item["statusCode"]) + " - " + item["statusText"])
        ws.cell(row=row_index, column=4, value=item["linkType"])
        ws.cell(row=row_index, column=5, value=item["fixGuide"])

        for col_index in range(1, 6):
            cell = ws.cell(row=row_index, column=col_index)
            cell.alignment = center_alignment if col_index == 3 or col_index == 4 else Alignment(vertical="center", wrap_text=True)
            cell.border = border

    # Adjust column widths
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = min(max_length + 5, 50)

    # Save to in-memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    try:
        time.tzset()
    except AttributeError:
        pass

    # Generate filename
    now_str = datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"Scan_Report_{scan_id}_{now_str}.xlsx"

    return filename, output 